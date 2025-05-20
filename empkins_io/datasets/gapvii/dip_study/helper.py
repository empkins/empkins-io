import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import Dict, Literal, Optional, Union, List
from biopsykit.utils.time import tz
from pandas import DataFrame
from empkins_io.utils._types import path_t
from empkins_io.sensors.emrad import EmradDataset
from empkins_io.sensors.tfm import TfmLoader
from empkins_io.sensors.empatica import EmpaticaDataset
import neurokit2 as nk
from scipy.signal import butter, filtfilt

import pandas as pd
import numpy as np

# Mapping German phase names to English phase keys used internally
PHASE_MAPPING = {
        "Beginn der Aufzeichnung": "start_recording",
        "Ende der Aufzeichnung": "end_recording",
        "Beginn Ruhe 1": "start_rest_1",
        "Ende Ruhe 1": "end_rest_1",
        "Beginn Ruhe 2": "start_rest_2",
        "Ende Ruhe 2": "end_rest_2",
        "Beginn Ruhe 3": "start_rest_3",
        "Ende Ruhe 3": "end_rest_3",
        "Beginn CPT": "start_cpt",
        "Ende CPT": "end_cpt",
        "Beginn Atmung": "start_straw",
        "Ende Atmung": "end_straw",
}

# Constants defining folder names for left and right device data storage
FOLDER_LEFT = "LEFT1-3YK33141NH"
FOLDER_RIGHT = "RIGHT-3YK34142G6"

# Prefixes for Empatica data files for left and right devices
PREFIX_LEFT = "1-1-LEFT1"
PREFIX_RIGHT = "1-1-RIGHT"

def _build_data_path(base_path: path_t, participant_id: str) -> Path:
    """
    Construct path to participant-specific data directory.
    """
    data_path = base_path.joinpath(f"data_per_subject/{participant_id}")
    assert data_path.exists()
    return data_path


def _build_tabular_data_path(base_path: path_t) -> Path:
    """
    Construct path to tabular data directory.
    """
    data_path = base_path.joinpath("data_tabular")
    assert data_path.exists()
    return data_path


def _build_general_tabular_path(base_path: path_t) -> Path:
    """
    Get path to the main processed Excel file with general tabular data.
    """
    data_path = _build_tabular_data_path(base_path)
    data_path = data_path.joinpath("processed/empkins_dip_study.xlsx")
    assert data_path.exists()
    return data_path

def _load_general_information(base_path: path_t, column: str) -> DataFrame:
    """
    Load a single column from the general tabular Excel file.
    Used to get participant metadata or other information.
    """
    file_path = _build_general_tabular_path(base_path)
    df = pd.read_excel(file_path, index_col=0)
    return df[column]

def _build_phase_times_path(base_path: path_t) -> Path:
    """
    Construct path to CSV file containing synchronized phase times.
    """
    data_path = _build_tabular_data_path(base_path)
    file_path = data_path / "processed/phase_times_synchronized.csv"
    assert file_path.exists(), f"{file_path} does not exist"
    return file_path

def _load_phase_times(base_path: path_t) -> DataFrame:
    """
    Load phase times CSV and convert start/end times from seconds to milliseconds.
    """
    file_path = _build_phase_times_path(base_path)
    df = pd.read_csv(file_path)

    # Convert start_time and end_time from seconds to milliseconds
    if "start_time" in df.columns and 'end_time' in df.columns:
        df["start_time"] = (df["start_time"] * 1000).astype("int64")
        df["end_time"] = (df["end_time"] * 1000).astype("int64")

    return df

def _build_datetime_path(base_path: path_t, participant_id: str) -> Path:
    """
    Build path to participant's cleaned protocol Excel file containing date info.
    """
    date_dir_path = _build_data_path(base_path, participant_id=participant_id).joinpath(
        "protocol/cleaned"
    )
    date_file_path = date_dir_path.joinpath(f"{participant_id}_protocol.xlsx")
    assert date_file_path.exists()
    return date_file_path

def _load_single_date(base_path: path_t, subject_id: str) -> pd.Timestamp:
    """
    Load the date (timestamp) from a specific cell (C3) in the "Allgemein" sheet of the protocol file.
    """
    data_path =_build_datetime_path(base_path=base_path, participant_id=subject_id)
    df = pd.read_excel(data_path, sheet_name="Allgemein", header=None)
    # Convert the cell to DataFrame indices (C3)
    cell_value = df.iloc[2, 2]
    return pd.to_datetime(cell_value, dayfirst=True)

def _update_dates(base_path: path_t, subject_date_dict: dict, sheet_name: str = "Sheet1"):
    """
    Update the 'date' column (H1) in the general tabular Excel file for given subjects.
    Writes formatted date string for each subject on matching row.
    """
    file_path = _build_general_tabular_path(base_path)
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet["H1"] = "date"

    # Loop through the names and update dates
    for subject, date in subject_date_dict.items():
        # Ensure the date is in the desired format
        if isinstance(date, pd.Timestamp):
            formatted_date = date.strftime("%d.%m.%Y")
        else:
            raise TypeError("Date should be a pd.Timestamp object")

        # Find the row with the matching name
        for row in sheet.iter_rows(min_row=2, max_row=None, min_col=1, max_col=1):
            cell = row[0]
            if cell.value == subject:
                # Write the date to the corresponding cell in column H
                sheet[f"H{cell.row}"] = formatted_date
                break

    # Save the changes to the workbook
    workbook.save(file_path)

def _load_start_end_times(base_path: path_t, participant_id: str) -> (datetime.datetime, datetime.datetime):
    """
    Load start and end times from specific cells (C39 and C175) in 'Messung' sheet
    of participant's protocol Excel file.
    """
    data_path = _build_data_path(base_path, participant_id=participant_id).joinpath(
        "protocol/cleaned"
    )
    data_path = data_path.joinpath(f"{participant_id}_protocol.xlsx")
    df = pd.read_excel(data_path, sheet_name="Messung", header=None)

    # Convert the cell to DataFrame indices (C39, C175)
    start_time = df.iloc[38, 2]
    end_time = df.iloc[174, 2]
    return (start_time, end_time)

def _create_loader(base_path: path_t, participant_id: str, date: str) -> TfmLoader:
    """
    Create a TfmLoader instance for the participant's TFM data at the given date.
    """
    tfm_dir_path = _build_data_path(base_path, participant_id=participant_id).joinpath(
        "tfm/cleaned"
    )

    # Convert the input date string to a datetime object using the specified format
    pd_date = pd.to_datetime(date, format='%d.%m.%Y')
    form_date = pd_date.strftime('%Y-%m-%d')

    # Construct the full path to the TFM data file for the participant
    tfm_file_path = tfm_dir_path.joinpath(f"{participant_id}_tfm_data.mat")

    # Load the TFM data from the specified .mat file
    loader = TfmLoader.from_mat_file(
        path=tfm_file_path,
        recording_date=form_date,  # Use the reformatted date
        phase_mapping=PHASE_MAPPING  # Map phases according to the dataset's mapping
    )

    return loader

def _load_tfm_data(base_path: path_t, participant_id: str, date: str) -> tuple[pd.DataFrame, float]:
    """
    Load TFM raw phase data and sampling rate for given participant and date.
    Returns data dictionary indexed by local_datetime and sampling frequency.
    """
    # Create a TfmLoader object for the specified participant and date
    loader = _create_loader(base_path, participant_id, date)

    # Extract the raw phase data as a dictionary of DataFrames, indexed by local datetime
    data = loader.raw_phase_data_as_df_dict(index="local_datetime")
    # Retrieve the sampling rates from the loader
    fs = loader.sampling_rates_hz

    return data, fs


def _load_b2b_data(base_path: path_t, participant_id: str, date: str) -> tuple[pd.DataFrame, float]:
    """
    Load beat-to-beat (B2B) data and sampling rate similarly to TFM data.
    """
    # Create a TfmLoader object for the specified participant and date
    loader = _create_loader(base_path, participant_id, date)
    # Extract the B2B phase data as a dictionary of DataFrames, indexed by local datetime
    data = loader.b2b_phase_data_as_df_dict(index="local_datetime")
    # Retrieve the sampling rates from the loader
    fs = loader.sampling_rates_hz
    return data, fs

def _load_radar_data(base_path: path_t, participant_id: str, sampling_rate_hz: float) -> tuple[pd.DataFrame, float]:
    """
    Load radar data from HDF5 file using EmradDataset class.
    Returns data as DataFrame indexed by local_datetime and sampling frequency.
    """
    # Build the directory path for radar data based on the base path and participant ID
    radar_dir_path = _build_data_path(base_path, participant_id=participant_id).joinpath(
        "emrad/raw"
    )
    
    # Build the file path to the radar data (HDF5 format) using the participant ID
    radar_file_path = radar_dir_path.joinpath(f"{participant_id}_emrad_data.h5")
    # Load radar data from the HDF5 file using the EmradDataset class
    dataset_radar = EmradDataset.from_hd5_file(radar_file_path, sampling_rate_hz=sampling_rate_hz)
    
    # Convert the radar data to a pandas DataFrame, using "local_datetime" as the index and adding a sync output column
    data = dataset_radar.data_as_df(index="local_datetime", add_sync_out=True)
    # Retrieve the sampling rate (Hz) from the dataset
    fs = dataset_radar.sampling_rate_hz
    # Return the DataFrame (data) and the sampling rate (fs)
    return data, fs

def _check_if_file_exists(base_path: path_t, subject_id: str, path_to_file) -> Optional[pd.DataFrame]:
    """
    Check if cached CSV file exists and load it.
    """
    data_path = _build_data_path(base_path, participant_id=subject_id)
    csv_path = data_path.joinpath(path_to_file)
    sampling_rates = {}

    # Load the existing file
    if csv_path.exists():
        df = pd.read_csv(csv_path)
        print(f"\tFile loaded from CSV Cache")
        return (df, sampling_rates)
    # If the file does not exist, return None
    else:
        return None

def _load_empatica_data(base_path: path_t, participant_id: str, date: str, empatica_lr: str, start_end_times: tuple[datetime.datetime, datetime.datetime], signal_type: list[str]):
    """
    Load Empatica data filtered by date, device side, signals, and time range.
    """
    data = {}
    
    # Convert date from "dd.mm.yyyy" to "yyyy-mm-dd" for folder matching
    date_parts = date.split(".")
    folder_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"

    # Determine LEFT or RIGHT device usage
    device_side = FOLDER_LEFT if empatica_lr == "L" else FOLDER_RIGHT

    # Determine the filename based on the device side, date, and signal type
    file_prefix = PREFIX_LEFT if empatica_lr == "L" else PREFIX_RIGHT

    for signal in signal_type:
        # Determine the filename based on the device side, date, and signal type
        filename = f"{file_prefix}_{folder_date}_{signal}.csv"

        # Build the path to the Empatica data directory for the given participant
        final_path = base_path.joinpath(f"Empatica/{folder_date}/{device_side}/digital_biomarkers/aggregated_per_minute/{filename}")

        # Check if file exists
        if not final_path.exists():
            print(f"File not found: {final_path}")
            continue

        # Load the Empatica data from the specified CSV file
        df = pd.read_csv(final_path, index_col=1, parse_dates=True)

        # Ensure index is in datetime format
        df.index = pd.to_datetime(df.index, utc=True)  # Ensure timestamps are timezone-aware in UTC

        # Convert UTC index to CEST (Europe/Berlin, GMT+2 in summer)
        df.index = df.index.tz_convert("Europe/Berlin")

        # Filter the data to the specified start and end times
        df = df.loc[start_end_times[0]:start_end_times[1]]

        # Append the filtered data to the dict
        data[signal] = df
        
    return data

def calculate_empatica_sampling_rate(df: pd.DataFrame) -> float:
    """
    Calculate average sampling rate from Unix timestamp differences in ms.
    """
    # Calculate the time difference between consecutive timestamps
    time_diff = df["timestamp_unix"].diff().dropna()
    # Calculate the average sampling rate in Hz
    avg_sampling_rate = 1 / (time_diff.mean() / 1000)  # Convert ms to seconds
    return avg_sampling_rate

def _create_agg_empatica(empatica_data: dict[str, pd.DataFrame], phase_times: pd.DataFrame) -> dict[str, dict[str, pd.DataFrame]]:
    """
    Aggregate Empatica signals per phase based on timestamp ranges.
    """
    empatica_data_by_phase = {}
    sampling_rates = {}

    # Iterate over each signal in the Empatica data
    for signal, df in empatica_data.items():
        # Ensure correct type
        df["timestamp_unix"] = pd.to_numeric(df["timestamp_unix"])
        df.columns.values[2] = "value"
        df["sampling_rate"] = calculate_empatica_sampling_rate(df)
        sampling_rates[f"{signal}_aggregated"] = df["sampling_rate"].iloc[0]

        phase_dict = {}
        # Iterate over each phase in the phase times DataFrame
        for _, row in phase_times.iterrows():
            phase = row["phase"]
            start = row["start_time"]
            end = row["end_time"]

            # Filter the DataFrame for the current phase
            sliced = df[
                (df["timestamp_unix"] >= start) &
                (df["timestamp_unix"] <= end)
            ]
            # Store the sliced DataFrame in the dictionary
            phase_dict[phase] = sliced

        # Store the aggregated data for the current signal
        empatica_data_by_phase[signal] = phase_dict
    return empatica_data_by_phase, sampling_rates

def _save_agg_empatica(base_path: path_t, subject_id: str, signal_phase_data: dict[str, dict[str, pd.DataFrame]], path_to_file: str) -> pd.DataFrame:
    """
    Save aggregated Empatica phase data into a single CSV file.
    """
    rows = []
    # Iterate over each signal and phase, and append the data to a list
    for signal, phases in signal_phase_data.items():
        for phase, df in phases.items():
            df_copy = df.copy()
            df_copy["timestamp"] = df_copy.index
            df_copy["signal"] = signal
            df_copy["phase"] = phase
            rows.append(df_copy)

    # Concatenate all rows into one big DataFrame
    full_df = pd.concat(rows, ignore_index=True)
    # order by timestamp
    full_df = full_df.sort_values(by=["signal","timestamp_unix"])

    # Reorder columns for clarity
    cols = ["signal", "phase", "timestamp"] + [col for col in full_df.columns if col not in ["signal", "phase", "timestamp"]]
    full_df = full_df[cols]

    # Save to CSV
    data_path = _build_data_path(base_path, participant_id=subject_id)
    output_path = data_path.joinpath(path_to_file)
    full_df.to_csv(output_path, index=False)
    print(f"\tSaved all data to: {output_path}")

    return full_df


def _bandpass_filter(signal, lowcut=0.5, highcut=8, fs=64, order=4):
    """
    Apply Butterworth bandpass filter to PPG signal.
        0.5 Hz = 30 bpm bellow physiological range
        8 Hz = 480 bpm above physiological range
        https://ieeexplore.ieee.org/document/9662889
    """
    # Normalize cutoff frequency by Nyquist frequency (fs/2)
    nyq = 0.5 * fs
    # Design filter (4rd order Butterworth)
    b, a = butter(order, [lowcut / nyq, highcut / nyq], btype="band")
    # Apply filter
    return filtfilt(b, a, signal)   



def _lowpass_filter(signal, cutoff=0.5, fs=4, order=3):
    """
    Apply Butterworth lowpass filter to EDA signal.
        filter out frequencies which are above 0.5 Hz
        https://www.sciencedirect.com/science/article/pii/S0167876021008461#bb0075
    """
    # Normalize cutoff frequency by Nyquist frequency (fs/2)
    normalized_cutoff = cutoff / (fs/2)
    # Design filter (3rd order Butterworth)
    b, a = butter(order, normalized_cutoff, 'low')
    # Apply filter
    return filtfilt(b, a, signal)

def _create_avro(base_path: path_t, participant_id: str, signal_type: list[str], phase_times: pd.DataFrame) -> tuple[dict[str, dict[str, pd.DataFrame]], float]:
    """
    Load and preprocess AVRO Empatica signals per phase with filtering and resampling.
    """
    # Build the path to the Empatica data directory for the given participant
    data_path = _build_data_path(base_path, participant_id=participant_id).joinpath("empatica/raw")

    # Load the Empatica data from the specified CSV file
    loader = EmpaticaDataset(path=data_path, index_type="local_datetime", tz="Europe/Berlin")
    sampling_rates = loader._sampling_rates_hz.copy()

    avro_data_by_phase = {}

    # Iterate over each signal type (e.g., "bvp", "eda")
    for signal in signal_type:
        # Load the data for the current signal
        df = loader.data_as_df(sensor=signal)

        # Ensure correct type
        df.columns.values[0] = "value"
        df.index.name = "timestamp"
        df["timestamp_unix"] = df.index.astype("int64") // 10**6
        df['sampling_rate'] = sampling_rates[signal]

        phase_dict = {}

        # Iterate over each phase in the phase times DataFrame
        for _, row in phase_times.iterrows():
            phase = row["phase"]
            start = row["start_time"]
            end = row["end_time"]

            # Filter the DataFrame for the current phase
            sliced = df[(df["timestamp_unix"] >= start) & (df["timestamp_unix"] <= end)].copy()
            
            # Check if the sliced DataFrame is empty
            if sliced.empty:
                print(f"No data for signal '{signal}' in phase '{phase}'")
                continue

            if signal == "bvp":
                fs = sampling_rates["bvp"]
                try:
                    # Bandpass filter
                    sliced["value"] = _bandpass_filter(sliced["value"], fs=fs)

                    # Process PPG
                    processed, info = nk.ppg_process(sliced["value"].values, sampling_rate=fs)
                    processed = processed.iloc[:len(sliced)]

                    # Add PPG Rate
                    sliced["value"] = processed["PPG_Rate"].values

                    # Ensure DatetimeIndex for resampling
                    if not isinstance(sliced.index, pd.DatetimeIndex):
                        sliced.index = pd.to_datetime(sliced.index)

                    # Downsample to 4 Hz (250ms)
                    new_sampling_rate = 4.0
                    sliced = sliced.resample("250ms").mean().interpolate("linear")
                    sliced["timestamp_unix"] = sliced.index.astype("int64") // 10**6
                    sliced["sampling_rate"] = new_sampling_rate

                except Exception as e:
                    print(f"Failed to process BVP for phase '{phase}' in subject '{participant_id}': {e}")
                    continue

            elif signal == "eda":
                fs = sampling_rates[signal]
                try:
                    # Apply lowpass filter
                    sliced["value"] = _lowpass_filter(sliced["value"].values, fs=fs)

                except Exception as e:
                    print(f"Failed to process EDA for phase '{phase}' in subject '{participant_id}': {e}")
                    continue

            phase_dict[phase] = sliced

        # Store the processed data for the current signal
        avro_data_by_phase[signal] = phase_dict

    return avro_data_by_phase, sampling_rates

def _save_avro(base_path: path_t, subject_id: str, signal_phase_data: dict[str, dict[str, pd.DataFrame]], path_to_file: str) -> pd.DataFrame:
    """
    Save processed AVRO Empatica data to CSV with phase and signal annotations.
    """
    rows = []
    # Iterate over each signal and phase, and append the data to a list
    for signal, phases in signal_phase_data.items():
        for phase, df in phases.items():
            df_copy = df.copy()
            df_copy["timestamp"] = df_copy.index
            df_copy["signal"] = signal
            df_copy["phase"] = phase
            rows.append(df_copy)

    # Concatenate all rows into one big DataFrame
    full_df = pd.concat(rows, ignore_index=True)
    # order by timestamp
    full_df = full_df.sort_values(by=["signal","timestamp_unix"])

    # Reorder columns for clarity
    cols = ["signal", "phase", "timestamp"] + [col for col in full_df.columns if col not in ["signal", "phase", "timestamp"]]
    full_df = full_df[cols]

    # Save to CSV
    data_path = _build_data_path(base_path, participant_id=subject_id)
    output_path = data_path.joinpath(path_to_file)
    full_df.to_csv(output_path, index=False)
    print(f"\tSaved all data to: {output_path}")

    return full_df

def _save_tfm_csv(base_path: path_t, subject_id: str, tfm_df: pd.DataFrame, path_to_file: str) -> pd.DataFrame:
    """
    Save TFM signal DataFrame to CSV, sorted and ordered by columns.
    """
    full_df = tfm_df.copy()
    
    # Sort rows by subject, signal, phase, timestamp_unix (timestamp_unix is numeric, so good for sorting)
    full_df = full_df.sort_values(by=["signal", "timestamp_unix"])

    # Reorder columns exactly as requested
    desired_order = [
        "signal", "phase", "timestamp", "timestamp_unix", 
        "value", "sampling_rate"
    ]

    # Keep only columns that exist in the DataFrame (in case some missing)
    existing_cols = [col for col in desired_order if col in full_df.columns]
    full_df = full_df[existing_cols]

    # Save to CSV
    data_path = _build_data_path(base_path, participant_id=subject_id)
    output_path = data_path.joinpath(path_to_file)
    full_df.to_csv(output_path, index=False)
    print(f"\tSaved all data to: {output_path}")

    return full_df
